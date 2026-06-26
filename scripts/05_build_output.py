#!/usr/bin/env python3
"""
05_build_output.py  --  Build final deliverables after curation.

Reads:
  04_output/curation/curated_annotations.csv   (all genes + needs_review flag)
  02_foldseek/3di_tokens/best_hit.csv          (for qcov_frac)
  input/overrides.tsv                          (per-gene manual overrides, applied here)
  the source GenBank / genome + gene metadata   (mode-dependent)

Writes:
  04_output/final_annotations_table.csv / .xlsx
  04_output/updated_prophages.gb
  04_output/curation/review_suggested.csv      (the genes flagged for a human look)
"""

import sys
from pathlib import Path

_SCRIPTS_DIR = Path(__file__).parent
_PROJECT_DIR = _SCRIPTS_DIR.parent
sys.path.insert(0, str(_PROJECT_DIR))

from config import (
    CURATION_DIR, OUTPUT_DIR, INPUT_DIR, GENE_METADATA_CSV, RAW_GB,
    FOLDSEEK_DIR, FOLDSEEK_BEST_HIT,
    FINAL_ANNOTATIONS_TABLE, FINAL_ANNOTATIONS_XLSX, UPDATED_GB,
    SOURCE_GENOME_FASTA, PROPHAGE_WINDOWS_CSV, PHOLD_AA_FASTA,
)
try:
    from config import RICH_GENE_METADATA_CSV
except ImportError:
    RICH_GENE_METADATA_CSV = None
from utils import log, section, clean_str, is_informative, build_note
try:
    from naming_rules import make_short_name
except Exception:
    def make_short_name(fp, phrog=None):  # graceful fallback
        return ""

# Genome-mode passthrough writes pharokka's PRODUCT into the function field
# (step 03 nulls phold's category for non-hypothetical genes), so final_function
# must be remapped to one of the 9 canonical PHROG categories. Self-contained
# keyword map (no fragile cross-module import) -> robust in every run.
import re as _re
_PHROG_CANON = {"head and packaging", "connector", "tail",
                "DNA, RNA and nucleotide metabolism", "integration and excision",
                "transcription regulation", "lysis",
                "moron, auxiliary metabolic gene and host takeover",
                "other", "unknown function"}
# ordered specific -> general; first hit wins
_CAT_KEYWORDS = [
    (_re.compile(r"integrase|recombinase|excisionase|transposase|resolvase|invertase|att(achment)? site", _re.I), "integration and excision"),
    (_re.compile(r"holin|endolysin|\blysin\b|lysozyme|spanin|\blysis\b|cell.wall (hydrolase|amidase)", _re.I), "lysis"),
    (_re.compile(r"terminase|portal|\bcapsid\b|prohead|procapsid|scaffold|head (protein|maturation|completion|decoration|closure)|major head|packaging", _re.I), "head and packaging"),
    (_re.compile(r"head.?tail|connector|\bneck\b|adaptor|adapter|stopper", _re.I), "connector"),
    (_re.compile(r"\btail\b|baseplate|tape.measure|fib(er|re)|spike|sheath|\btube\b|whisker|virion structural|distal tail|Dit\b|tail terminator", _re.I), "tail"),
    (_re.compile(r"helicase|polymerase|primase|exonuclease|endonuclease|nuclease|\bkinase\b|methyltransferase|methylase|\bligase\b|replicati|topoisomerase|single.strand|ssDNA|ssb\b|nucleotide|ribonucleotide|DNA.binding|recombination|annealing|Holliday", _re.I), "DNA, RNA and nucleotide metabolism"),
    (_re.compile(r"repressor|regulator|\bcro\b|\bci\b|antiterminat|anti.terminat|transcription|sigma.factor|\bHTH\b|anti.repressor|Cox\b", _re.I), "transcription regulation"),
    (_re.compile(r"defen[cs]e|restriction|toxin|antitoxin|abortive infection|anti.crispr|\bCBASS\b|\bPARIS\b|\bRM\b system|immunity", _re.I), "defense"),
    (_re.compile(r"ribosom|\bzur\b|fur.family|oxidoreductase|dehydrogenase|reductase|transferase|permease|metaboli|tRNA|amino.acid|sugar|transport|hydrolase|phosphatase|synthase|isomerase|racemase|epimerase|moron", _re.I), "moron, auxiliary metabolic gene and host takeover"),
]

def _phrog_category(function, product):
    """Return a canonical PHROG category. Keep `function` if already canonical,
    else infer from the product via the keyword map; else other/unknown."""
    f = str(function or "").strip()
    if f in _PHROG_CANON:
        return f
    text = f"{product or ''} {f}"
    for pat, cat in _CAT_KEYWORDS:
        if pat.search(text):
            return cat
    return "other" if f and f.lower() not in ("", "nan", "na") else "unknown function"

try:
    from Bio import SeqIO
    import pandas as pd
except ImportError as e:
    print(f"Missing dependency: {e}"); sys.exit(1)


# ---------------------------------------------------------------------------
# HELPERS
# ---------------------------------------------------------------------------

def _safe_float(v):
    import math
    try:
        f = float(v)
        return f if not math.isnan(f) else None
    except (TypeError, ValueError):
        return None


def _bool_flag(v) -> bool:
    if v is None: return False
    if isinstance(v, float) and pd.isna(v): return False
    if isinstance(v, bool): return v
    if isinstance(v, (int, float)): return bool(v)
    s = str(v).strip().lower()
    return s in ("true", "1", "yes")


def _find_curation_files(curation_dir: Path):
    # 2026-06-26: curated files now live in 04_output/curation/; also look in the
    # parent (old 04_output/ location) so previously-generated runs still load.
    search_dirs = [curation_dir, curation_dir.parent]
    auto_path = None
    # 2026-06: prefer the combined curated_annotations.csv (all genes, one file);
    # fall back to the legacy auto_curated.csv split for older runs.
    for d in search_dirs:
        for name in ("curated_annotations.csv", "auto_curated.csv", "auto_curated_new.csv"):
            p = d / name
            if p.exists(): auto_path = p; break
        if auto_path: break
    review_path = None
    # review_suggested.csv is the new name; keep old names for backward compat
    for d in search_dirs:
        for name in ("review_suggested.csv", "needs_review.csv", "needs_review_NEW.csv", "needs_review_new.csv"):
            p = d / name
            if p.exists():
                try:
                    open(str(p)).close()   # verify readable (NTFS case-insensitive exists() can lie)
                    review_path = p; break
                except OSError:
                    continue
        if review_path: break
    return auto_path, review_path


def _load_curated_annotations(curation_dir: Path) -> pd.DataFrame:
    auto_path, review_path = _find_curation_files(curation_dir)
    dfs = []
    # When the single combined file is present it already contains the review rows,
    # so we must NOT also append review_suggested.csv (would double-count).
    _combined = bool(auto_path and auto_path.name == "curated_annotations.csv")

    if auto_path:
        auto = pd.read_csv(str(auto_path))
        if "final_annotation" not in auto.columns:
            auto["final_annotation"] = auto.get("final_product", "hypothetical protein")
        src = auto.get("best_source", pd.Series("unknown", index=auto.index)).fillna("unknown")
        src = src.replace({"none": "no_hit", "NA": "no_hit", "nan": "no_hit", "both_uninformative": "no_hit"})
        auto["annotation_source"] = src
        # In the combined file, flag user-edited review rows as manual_review.
        if _combined and "needs_review" in auto.columns:
            flagged = auto["needs_review"] == True
            edited = flagged & (auto.get("final_annotation", "").astype(str).str.strip()
                                != auto.get("AI_suggestion", "").astype(str).str.strip())
            auto.loc[edited, "annotation_source"] = "manual_review"
        dfs.append(auto)
        log(f"  {auto_path.name}: {len(auto)} rows" + ("  [combined]" if _combined else ""))
    else:
        log("WARNING: no curated CSV found in 04_curation/")

    if review_path and not _combined:
        review = pd.read_csv(str(review_path))
        n_total = len(review)
        # Fill any blank final_annotation with AI_suggestion
        for idx, row in review.iterrows():
            if not isinstance(row.get("final_annotation"), str) or \
               not str(row["final_annotation"]).strip():
                review.at[idx, "final_annotation"] = row.get("AI_suggestion", "hypothetical protein")
        n_filled = review["final_annotation"].apply(
            lambda x: isinstance(x, str) and len(x.strip()) > 0).sum()
        log(f"  {review_path.name}: {len(review)} rows ({n_filled} final_annotation filled)")

        # Attribution rule (fixed 2026-05-27):
        # For review rows, preserve the underlying evidence source from best_source
        # (set by 04_curate_annotations.py: phold / foldseek / phold+foldseek / none).
        # Only fall back to "manual_review" when the user typed a different name,
        # or "AI_suggestion" if best_source is missing/uninformative.
        # 2026-06: 04_curate now emits "both agree" / "merged" instead of "phold+foldseek"
        # (older value kept for backward compat with previously-generated tables).
        VALID_SOURCES = {"phold", "foldseek", "both agree", "merged", "phold+foldseek", "no_hit", "none"}

        def _review_source(row):
            fa = str(row.get("final_annotation", "")).strip()
            ai = str(row.get("AI_suggestion",    "")).strip()
            bs = str(row.get("best_source",      "")).strip().lower()
            # User typed a name different from the AI suggestion -> manual override
            if fa and fa != ai:
                return "manual_review"
            # Accepted suggestion: attribute to the underlying evidence source
            if bs in VALID_SOURCES and bs not in ("no_hit", "none", "flagged"):
                return bs
            # No usable evidence in best_source -> fallback (rare with new Case 5)
            return "AI_suggestion"

        review["annotation_source"] = review.apply(_review_source, axis=1)
        dfs.append(review)
    else:
        log("INFO: no needs_review CSV found -- all genes auto-curated.")

    if not dfs:
        log("ERROR: no curation files found."); sys.exit(1)

    combined = pd.concat(dfs, ignore_index=True)
    log(f"  Total curated: {len(combined)} hypothetical genes")
    return combined


def _apply_overrides(curated: pd.DataFrame, overrides_path: Path) -> pd.DataFrame:
    """Apply hand-curated per-gene overrides (input/overrides.tsv) to the curated
    table BEFORE the final table + review_suggested are built, so an override both
    fixes the call AND clears the review flag (07_integrate re-applies the same
    file to the phynteny-integrated table, idempotently). TSV columns:
      match_type (locus_tag|final_product), match, final_product, final_function,
      short_name, annotation_source, note  -- any blank field leaves it unchanged.
    An override always sets needs_review=False for the matched rows."""
    if not overrides_path or not overrides_path.exists():
        return curated
    ov = pd.read_csv(str(overrides_path), sep="\t", dtype=str, keep_default_na=False)
    if "short_name" not in curated.columns:
        curated["short_name"] = ""
    n = 0
    for _, o in ov.iterrows():
        mt, mv = str(o.get("match_type", "")).strip(), str(o.get("match", "")).strip()
        if not mt or not mv:
            continue
        col = "locus_tag" if mt == "locus_tag" else "final_product"
        if col not in curated.columns:
            continue
        sel = curated[col].astype(str) == mv
        if not sel.any():
            continue
        fp = str(o.get("final_product", "")).strip()
        if fp:
            for c in ("final_product", "final_annotation"):
                if c in curated.columns:
                    curated.loc[sel, c] = fp
        for src_col, dst_col in (("final_function", "final_function"),
                                 ("short_name", "short_name"),
                                 ("annotation_source", "annotation_source"),
                                 ("annotation_source", "best_source")):
            val = str(o.get(src_col, "")).strip()
            if val and dst_col in curated.columns:
                curated.loc[sel, dst_col] = val
        if "needs_review" in curated.columns:
            curated.loc[sel, "needs_review"] = False
        n += int(sel.sum())
    if n:
        log(f"  Applied {n} override row(s) from {overrides_path.name} (flags cleared)")
    return curated


def _load_foldseek_qcov(foldseek_dir: Path) -> dict:
    # Use FOLDSEEK_BEST_HIT from config (cluster: 3di_tokens/best_hit.csv;
    # local: foldseek_best_hit.csv) rather than constructing the path here.
    best_hit = FOLDSEEK_BEST_HIT
    if not best_hit.exists():
        # Fallback to legacy local path in case config doesn't define it
        best_hit = foldseek_dir / "foldseek_best_hit.csv"
    if not best_hit.exists():
        log(f"WARNING: {best_hit} not found -- qcov_frac/partial_match will be NA")
        return {}
    df = pd.read_csv(str(best_hit))
    out = {}
    for _, row in df.iterrows():
        gene = str(row.get("gene", "")).strip()
        if gene:
            out[gene] = {
                "foldseek_qcov_frac":     _safe_float(row.get("foldseek_qcov_frac")),
                "foldseek_partial_match": _bool_flag(row.get("foldseek_partial_match", False)),
            }
    return out


# ---------------------------------------------------------------------------
# GENBANK UPDATER
# ---------------------------------------------------------------------------

def update_gb(original_gb_path: Path, annotation_map: dict, output_gb_path: Path) -> int:
    records = list(SeqIO.parse(str(original_gb_path), "genbank"))
    n_updated = 0
    for record in records:
        for feature in record.features:
            if feature.type != "CDS": continue
            lt = feature.qualifiers.get("locus_tag", [""])[0].strip()
            if lt not in annotation_map: continue
            ann = annotation_map[lt]
            feature.qualifiers["product"]  = [ann.get("product",  "hypothetical protein")]
            feature.qualifiers["function"] = [ann.get("function", "unknown function")]
            note_text = ann.get("note", "")
            if note_text:
                existing = feature.qualifiers.get("note", [])
                if not any("pipeline=phagefactor" in n for n in existing):
                    feature.qualifiers["note"] = [note_text] + existing
            n_updated += 1
    output_gb_path.parent.mkdir(parents=True, exist_ok=True)
    with open(str(output_gb_path), "w") as fh:
        SeqIO.write(records, fh, "genbank")
    return n_updated


# ---------------------------------------------------------------------------
# GENBANK BUILDER (pre-CDS-computed source-genome submode -- 2026-06, citro)
#
# For datasets like citro (phiNP/phiSM): proteins came from a pre-annotated
# source genome's precomputed CDS (eggNOG), via protein mode -- so there is
# no Pharokka GBK for update_gb() to update. If gene_metadata.csv carries
# ABSOLUTE host-genome coordinates (the "pre-CDS-computed" submode, as
# opposed to plain pangenome protein mode which has none), and the user
# supplies the source genome FASTA + prophage window coordinates, this
# builds a base per-prophage GBK directly -- which update_gb() then overlays
# with curated final annotations exactly as it does for genome-mode runs.
#
# ID design (per user decision -- "carry over the original source genome
# gene IDs", confirmed against the DE-join requirement, see chat 2026-06-07):
#   locus_tag     = gene_id_fa   (pipeline-internal `fig|NNNNNXNNXpegXN`,
#                                 matches final_annotations_table.csv's
#                                 locus_tag -- so update_gb()/annotation_map
#                                 keys land on the right CDS automatically)
#   old_locus_tag = gene_id_gff  (original source-genome dotted ID
#                                 `fig|NNNNN.NN.peg.N` -- preserved for
#                                 provenance / to satisfy the "carry over
#                                 original IDs" decision; this is also
#                                 exactly the format de_results/defense_lfc
#                                 use as their join key, see chat assessment)
# ---------------------------------------------------------------------------

def _load_prophage_windows(path: Path) -> dict:
    """Read PROPHAGE_WINDOWS_CSV -> {prophage: (start, stop, contig)}.
    `contig` (optional 4th column) lets the GBK builder pick the right contig when
    prophages span MULTIPLE host contigs (e.g. several prophages each on a different
    bacterium; a contig may carry >1 prophage = polylysogen). contig=None if absent."""
    if not path.exists():
        return {}
    df = pd.read_csv(str(path))
    has_contig = "contig" in df.columns
    out = {}
    for _, r in df.iterrows():
        try:
            contig = str(r["contig"]).strip() if has_contig and pd.notna(r.get("contig")) else None
            out[str(r["prophage"]).strip()] = (int(r["genome_start"]), int(r["genome_stop"]), contig)
        except (KeyError, ValueError, TypeError):
            continue
    return out


def _load_aa_sequences(aa_fasta: Path) -> dict:
    """Map bare locus_tag (after stripping a 'PROPHAGE:' prefix) -> AA sequence string."""
    seqs = {}
    if not aa_fasta.exists():
        return seqs
    cur, buf = None, []
    with open(str(aa_fasta)) as fh:
        for line in fh:
            line = line.rstrip()
            if line.startswith(">"):
                if cur is not None:
                    seqs[cur] = "".join(buf)
                hdr = line[1:].split()[0]
                cur = hdr.split(":", 1)[1] if ":" in hdr else hdr
                buf = []
            else:
                buf.append(line.strip())
        if cur is not None:
            seqs[cur] = "".join(buf)
    return seqs


def _enrich_metadata_for_gbk(metadata: "pd.DataFrame", rich_csv: Path) -> "pd.DataFrame":
    """
    Bridge the schema gap between protein-mode's MINIMAL gene_metadata.csv
    (prophage,locus_tag,function,aa_length,is_hypothetical -- regenerated by
    01p_merge_phold_proteins._write_gene_metadata, no per-gene Pharokka GFF
    to source product/coords from) and what build_gbk_from_metadata() needs
    (absolute host-genome coords: gene_id_gff,start,end,strand,gff_product,
    gene_id_fa). The richer columns live ONLY in the ORIGINAL pre-CDS-computed
    annotation metadata (e.g. Citro_prophage_annotation/input/gene_metadata.csv
    for citro -- schema: prophage,gene_id_gff,gene_id_fa,start,end,strand,
    aa_length,gff_product,emapper_desc,...).

    Joins on locus_tag (split metadata) == gene_id_fa (rich metadata) -- both
    are the `fig|...peg...` protein IDs and are globally unique, so a left
    merge is safe. Returns a NEW enriched DataFrame for GBK building only;
    the original `metadata` (passed to everything else in main()) is left
    untouched so the rest of the pipeline keeps using the minimal schema it
    expects. Returns the original `metadata` unchanged if the rich file is
    missing/unreadable or doesn't actually carry the needed columns (caller's
    existing needed_cols guard then logs the familiar "skip" message).
    """
    if rich_csv is None or not Path(rich_csv).exists():
        log(f"  No rich gene metadata at {rich_csv} -- GBK build will use the "
            "minimal schema (likely skipped, see needed_cols check below)")
        return metadata

    try:
        rich = pd.read_csv(str(rich_csv))
    except Exception as exc:
        log(f"  WARNING: could not read rich gene metadata {rich_csv}: {exc}")
        return metadata

    needed = {"gene_id_fa", "gene_id_gff", "start", "end", "strand"}
    if not needed.issubset(rich.columns):
        log(f"  WARNING: rich gene metadata {rich_csv} missing required columns "
            f"(has: {sorted(rich.columns)}, need: {sorted(needed)}) -- ignoring it")
        return metadata

    cols = ["gene_id_fa", "gene_id_gff", "start", "end", "strand"]
    if "gff_product" in rich.columns:
        cols.append("gff_product")
    rich_sub = rich[cols].drop_duplicates(subset="gene_id_fa")

    # Drop any of these columns already present in `metadata` BEFORE the merge,
    # else pandas suffixes them start_x/start_y/... and build_gbk_from_metadata
    # can't find the plain start/end/strand/gene_id_gff (the bug that silently
    # skipped the GBK when the minimal metadata already carried coords). rich is
    # the authoritative source for these.
    meta_clean = metadata.drop(columns=[c for c in cols if c in metadata.columns],
                               errors="ignore")
    enriched = meta_clean.merge(rich_sub, how="left",
                                left_on="locus_tag", right_on="gene_id_fa")
    n_matched = enriched["gene_id_gff"].notna().sum()
    log(f"  Enriched gene metadata for GBK from {rich_csv.name}: "
        f"{n_matched}/{len(enriched)} genes matched (locus_tag == gene_id_fa)")
    return enriched


def build_gbk_from_metadata(metadata: "pd.DataFrame", source_fasta: Path,
                            windows_csv: Path, aa_fasta: Path,
                            out_path: Path) -> int:
    """
    Build a base (pre-curation) multi-record GenBank for the pre-CDS-computed
    submode (gene_metadata.csv carries absolute host-genome coordinates).
    Returns the number of CDS features written (0 if prerequisites are missing
    -- caller treats that as "skip, fall through to existing warning").
    """
    from Bio.SeqRecord import SeqRecord
    from Bio.SeqFeature import SeqFeature, FeatureLocation

    needed_cols = {"prophage", "gene_id_gff", "start", "end", "strand"}
    lt_col = "gene_id_fa" if "gene_id_fa" in metadata.columns else (
        "locus_tag" if "locus_tag" in metadata.columns else None)
    if lt_col is None or not needed_cols.issubset(metadata.columns):
        log(f"  Not the pre-CDS-computed submode (gene_metadata.csv missing absolute "
            f"coords / gene_id_gff / gene_id_fa columns -- has: {sorted(metadata.columns)}) "
            "-- skipping metadata-based GBK build")
        return 0

    if not source_fasta.exists():
        log(f"  Source genome FASTA not found: {source_fasta}")
        return 0

    windows = _load_prophage_windows(windows_csv)
    if not windows:
        log(f"  No prophage window coordinates at {windows_csv} "
            "(expected columns: prophage,genome_start,genome_stop)")
        return 0

    genome_records = list(SeqIO.parse(str(source_fasta), "fasta"))
    if not genome_records:
        log(f"  Source genome FASTA is empty: {source_fasta}")
        return 0
    # Multi-contig aware: each prophage's window is sliced from ITS OWN contig
    # (a multi-bacterium input can have several prophages across distinct contigs).
    genome_dict = {r.id.split()[0]: r.seq for r in genome_records}

    aa_seqs = _load_aa_sequences(aa_fasta)
    log(f"  Loaded {len(aa_seqs)} AA sequences from {aa_fasta.name} for /translation")

    records = []
    n_cds = 0
    for prophage, (gstart, gstop, contig) in windows.items():
        sub = metadata[metadata["prophage"] == prophage]
        if sub.empty:
            log(f"  WARNING: no gene_metadata rows for window '{prophage}' -- skipping record")
            continue

        # pick this prophage's contig (fall back to first contig if unspecified)
        if contig and contig in genome_dict:
            cseq = genome_dict[contig]
        else:
            cseq = genome_records[0].seq
            if contig:
                log(f"  WARNING: contig '{contig}' for {prophage} not in source FASTA "
                    f"-- using first contig {genome_records[0].id}")
        clen = len(cseq)
        wrap = gstop < gstart            # window spans the (circular) genome origin
        rec_seq = (cseq[gstart - 1:] + cseq[:gstop]) if wrap else cseq[gstart - 1:gstop]

        features = []
        for _, g in sub.sort_values("start").iterrows():
            try:
                abs_start, abs_end = int(g["start"]), int(g["end"])
            except (TypeError, ValueError):
                continue
            if wrap:
                # offset within the concatenated [gstart..end]+[1..gstop] record
                rel_start = (abs_start - gstart) % clen
                rel_end   = (abs_end - gstart) % clen
            else:
                rel_start, rel_end = abs_start - gstart, abs_end - gstart
            if rel_start < 0 or rel_end > len(rec_seq) or rel_start >= rel_end:
                log(f"  WARNING: {g[lt_col]} coords ({abs_start}-{abs_end}) fall outside "
                    f"window {prophage} ({gstart}-{gstop}) -- skipping feature")
                continue
            strand = -1 if str(g["strand"]).strip() in ("-", "-1") else 1
            lt  = str(g[lt_col]).strip()
            old = str(g["gene_id_gff"]).strip()
            qualifiers = {
                "locus_tag":     [lt],
                "old_locus_tag": [old],
                "product":       [str(g.get("gff_product") or "hypothetical protein")],
            }
            translation = aa_seqs.get(lt) or aa_seqs.get(old)
            if translation:
                qualifiers["translation"]  = [translation]
                qualifiers["transl_table"] = ["11"]
            features.append(SeqFeature(FeatureLocation(rel_start, rel_end, strand=strand),
                                        type="CDS", qualifiers=qualifiers))
            n_cds += 1

        rec = SeqRecord(rec_seq, id=prophage, name=prophage[:16],
                        description=f"{prophage} prophage region "
                                    f"({gstart}-{gstop} on contig {contig or genome_records[0].id})",
                        features=features)
        rec.annotations["molecule_type"] = "DNA"
        records.append(rec)
        log(f"  Built {prophage}: {len(rec_seq)} bp, {len(features)} CDS  (window {gstart}-{gstop})")

    if not records:
        return 0

    out_path.parent.mkdir(parents=True, exist_ok=True)
    with open(str(out_path), "w") as fh:
        SeqIO.write(records, fh, "genbank")
    log(f"  Wrote base GBK -> {out_path}  ({len(records)} records, {n_cds} CDS)")
    return n_cds


# ---------------------------------------------------------------------------
# EXCEL WRITER
# ---------------------------------------------------------------------------

def _write_excel(df: pd.DataFrame, path: Path):
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    HEADER_FILL   = PatternFill("solid", fgColor="2B5797")
    RESOLVED_FILL = PatternFill("solid", fgColor="EAF4E8")
    STILL_FILL    = PatternFill("solid", fgColor="FCE4D6")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Final Annotations"

    headers = list(df.columns)
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"

    for ri, (_, row) in enumerate(df.iterrows(), 2):
        is_hypo    = _bool_flag(row.get("was_hypothetical", False))
        still_hypo = str(row.get("final_product", "")) == "hypothetical protein"
        fill = (STILL_FILL if (is_hypo and still_hypo) else
                RESOLVED_FILL if is_hypo else None)
        for ci, h in enumerate(headers, 1):
            val = row[h]
            if hasattr(val, "item"): val = val.item()
            cell = ws.cell(row=ri, column=ci, value=val)
            if fill: cell.fill = fill

    for ci, h in enumerate(headers, 1):
        col_letter = get_column_letter(ci)
        max_w = max(len(str(h)),
                    *(min(len(str(ws.cell(r, ci).value or "")), 60)
                      for r in range(2, min(ws.max_row + 1, 302))))
        ws.column_dimensions[col_letter].width = min(max_w + 2, 60)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------

def main():
    section("STEP 05 -- BUILD FINAL OUTPUTS")
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    if not GENE_METADATA_CSV.exists():
        log(f"ERROR: {GENE_METADATA_CSV} not found."); sys.exit(1)
    metadata = pd.read_csv(str(GENE_METADATA_CSV))
    log(f"Gene metadata: {len(metadata)} genes")

    section("LOADING CURATED ANNOTATIONS")
    curated = _load_curated_annotations(CURATION_DIR)
    curated = _apply_overrides(curated, INPUT_DIR / "overrides.tsv")
    curated_idx = curated.set_index("locus_tag")

    qcov_lookup = _load_foldseek_qcov(FOLDSEEK_DIR)

    section("BUILDING FINAL ANNOTATION TABLE")
    rows = []
    annotation_map = {}

    for _, meta_row in metadata.iterrows():
        lt            = meta_row["locus_tag"]
        prophage      = meta_row["prophage"]
        is_hypo       = meta_row["is_hypothetical"]
        # protein-mode gene_metadata.csv (regenerated by 01p_merge_phold_proteins.
        # _write_gene_metadata) is intentionally minimal -- it only carries
        # prophage/locus_tag/function/aa_length/is_hypothetical (no product/
        # start/end/strand, since there is no Pharokka GFF for these genes;
        # "All input proteins treated as hypothetical candidates" per
        # submit_all.sh). Genome-mode metadata DOES carry "product" (from
        # parse_gb_to_metadata), so use .get() with the semantically-correct
        # fallback rather than a hard KeyError -- a missing original product
        # IS "hypothetical protein" for these pre-CDS-computed datasets (e.g.
        # citro), exactly mirroring is_hypothetical=True for every row.
        orig_product  = meta_row.get("product",  "hypothetical protein")
        orig_function = meta_row.get("function", "unknown function")
        aa_len        = meta_row["aa_length"]

        if is_hypo and lt in curated_idx.index:
            cur = curated_idx.loc[lt]

            def _get(col, fallback="NA"):
                v = cur[col] if col in cur.index else fallback
                if hasattr(v, "iloc"): v = v.iloc[0]
                return v

            final_product  = clean_str(_get("final_annotation", "hypothetical protein"))
            if not is_informative(final_product): final_product = "hypothetical protein"
            final_function = clean_str(_get("final_function", "unknown function"))
            if not is_informative(final_function): final_function = "unknown function"

            annotation_source = clean_str(_get("annotation_source", "unknown"))
            action            = clean_str(_get("curation_action",   "NA"))

            # PHold evidence columns (now exposed in final table for traceability)
            p_product = clean_str(_get("phold_product"))
            p_cat     = clean_str(_get("phold_function_cat"))
            p_conf    = clean_str(_get("phold_confidence"))
            p_evalue  = _safe_float(_get("phold_evalue"))
            p_phrog   = clean_str(_get("phold_phrog"))
            # accession_phrog == phold_phrog (same PHROG accession; carried
            # through under both names per user request -- "phold phrog is in
            # phold_phrog (for accession_phrog if needed)")
            accession_phrog = p_phrog

            # agreement: produced by 03_compare, carried verbatim through
            # 04_curate's output -- "agreement is in both documents the same
            # in the step04 output => just make sure to add [it] to step05"
            agreement = clean_str(_get("agreement", "NA"))

            # FoldSeek evidence columns (also exposed for traceability)
            fs_desc      = clean_str(_get("foldseek_description"))
            fs_acc       = clean_str(_get("foldseek_accession"))
            fs_db        = clean_str(_get("foldseek_db"))         # source DB: afdb-swissprot / pdb100 / afdb50
            fs_taxname   = clean_str(_get("foldseek_taxname"))
            fs_kingdom   = clean_str(_get("best_hit_kingdom"))   # filled by 03 (eukaryotic filter)
            fs_conf      = clean_str(_get("foldseek_confidence"))
            fs_eval      = _safe_float(_get("foldseek_evalue"))
            fs_score     = _safe_float(_get("foldseek_score"))
            fs_pid       = _safe_float(_get("foldseek_pident"))
            fs_top3      = clean_str(_get("foldseek_top3"))
            fs_top3_king = clean_str(_get("foldseek_top3_kingdoms"))  # kingdom per top-3 hit
            fs_same_host = _bool_flag(_get("foldseek_same_host", False))
            sn_override  = clean_str(_get("short_name", ""))          # from input/overrides.tsv

            qd = qcov_lookup.get(lt, {})
            fs_qcov_frac     = qd.get("foldseek_qcov_frac")
            fs_partial_match = qd.get("foldseek_partial_match", False)

            note_text = build_note(
                source=annotation_source,
                phold_confidence=p_conf if p_conf not in ("NA","none","nan") else None,
                phold_evalue=p_evalue, phold_phrog=p_phrog,
                foldseek_evalue=fs_eval, foldseek_score=fs_score, foldseek_pident=fs_pid,
            )
            annotation_map[lt] = {"product": final_product, "function": final_function, "note": note_text}

        else:
            final_product = orig_product
            final_function = _phrog_category(orig_function, orig_product)
            annotation_source = "pharokka"; action = "kept_pharokka"
            p_product = p_cat = p_conf = p_phrog = None
            accession_phrog = None
            agreement = "NA"
            p_evalue  = None
            fs_desc = fs_acc = fs_db = fs_taxname = fs_kingdom = None
            fs_conf = fs_eval = fs_score = fs_pid = fs_top3 = fs_same_host = None
            fs_top3_king = None
            fs_qcov_frac = fs_partial_match = None
            sn_override = ""
            note_text = ""

        rows.append({
            # ---- Identity ----
            "prophage":                 prophage,
            "locus_tag":                lt,
            "aa_length":                aa_len,
            # ---- Curated best call (early so reviewers see the answer first) ----
            "short_name":               (sn_override if is_informative(sn_override)
                                         else make_short_name(final_product, p_phrog if is_hypo else None)),
            "final_product":            final_product,
            "final_function":           final_function,
            "was_hypothetical":         is_hypo,
            "annotation_source":        annotation_source,
            "best_hit_kingdom":         fs_kingdom if is_hypo else "NA",
            # ---- Pharokka columns ----
            "pharokka_product":         orig_product,
            "pharokka_function":        orig_function,
            # ---- Phold columns ----
            "phold_product":            p_product if is_hypo else "NA",
            "phold_function_cat":       p_cat     if is_hypo else "NA",
            "phold_confidence":         p_conf    if is_hypo else "NA",
            "phold_phrog":              p_phrog   if is_hypo else "NA",
            "accession_phrog":          accession_phrog if is_hypo else "NA",
            "phold_evalue":             p_evalue,
            # ---- Curation evidence agreement (carried verbatim from step04) ----
            "agreement":                agreement if is_hypo else "NA",
            "curation_action":          action if is_hypo else "NA",
            "needs_review":             bool(_get("needs_review", False)) if is_hypo else False,
            # ---- Custom FoldSeek columns (fs_* for readability) ----
            "fs_description":     fs_desc    if is_hypo else "NA",
            "fs_accession":       fs_acc     if is_hypo else "NA",
            "fs_db":              fs_db      if is_hypo else "NA",
            "fs_taxname":         fs_taxname if is_hypo else "NA",
            "fs_confidence":      fs_conf    if is_hypo else "NA",
            "fs_score":           fs_score,
            "fs_evalue":          fs_eval,
            "fs_pident":          fs_pid,
            "fs_qcov_frac":       fs_qcov_frac,
            "fs_same_host":       fs_same_host,
            "fs_top3":            fs_top3,
            "fs_top3_kingdoms":   fs_top3_king if is_hypo else "NA",
            # ---- Provenance ----
            "note":                     note_text,
        })

    final_df = pd.DataFrame(rows).sort_values(["prophage", "locus_tag"])

    FINAL_ANNOTATIONS_TABLE.parent.mkdir(parents=True, exist_ok=True)
    final_df.to_csv(str(FINAL_ANNOTATIONS_TABLE), index=False)
    log(f"CSV  -> {FINAL_ANNOTATIONS_TABLE}  ({len(final_df)} rows, {len(final_df.columns)} cols)")

    # 2026-06: emit the small review subset alongside the final table, so the
    # genes still needing a human look live next to the deliverable (step04's split
    # is now merged into one curated file). Empty file written if nothing to review.
    review_view = final_df[final_df.get("needs_review", False) == True].copy()
    CURATION_DIR.mkdir(parents=True, exist_ok=True)
    review_out = CURATION_DIR / "review_suggested.csv"
    review_view.to_csv(str(review_out), index=False)
    log(f"CSV  -> {review_out}  ({len(review_view)} flagged for review)")

    try:
        import openpyxl
        _write_excel(final_df, FINAL_ANNOTATIONS_XLSX)
        log(f"XLSX -> {FINAL_ANNOTATIONS_XLSX}")
    except ImportError:
        log("WARNING: openpyxl not installed -- skipping Excel output")

    # -- Statistics -----------------------------------------------------------
    section("FINAL ANNOTATION STATISTICS")
    total      = len(final_df)
    hypo_mask  = final_df["was_hypothetical"]
    hypo_total = hypo_mask.sum()
    non_hypo   = total - hypo_total

    annotated_mask = hypo_mask & final_df["final_product"].apply(is_informative)
    now_annotated  = annotated_mask.sum()
    still_hypo     = hypo_total - now_annotated

    log(f"Total CDS                  : {total}")
    log(f"  Non-hypothetical          : {non_hypo}  (Pharokka kept)")
    log(f"  Originally hypothetical   : {hypo_total}")
    log(f"    Resolved by pipeline    : {now_annotated}  ({100*now_annotated//hypo_total}%)")
    log(f"    Still hypothetical      : {still_hypo}")
    log(f"\nOverall annotation coverage : {non_hypo+now_annotated}/{total}  ({100*(non_hypo+now_annotated)//total}%)")

    log("\nAnnotation sources (hypothetical genes):")
    log(final_df[hypo_mask]["annotation_source"].value_counts().to_string())

    log("\nFoldSeek confidence tiers (hypothetical genes):")
    log(final_df[hypo_mask]["fs_confidence"].value_counts().to_string())

    log("\nPer-prophage resolution:")
    try:
        pp = final_df.groupby("prophage").apply(
            lambda g: pd.Series({
                "total":         len(g),
                "hypo_in":       g["was_hypothetical"].sum(),
                "resolved":      (g["was_hypothetical"] & g["final_product"].apply(is_informative)).sum(),
                "still_hypo":    (g["was_hypothetical"] & ~g["final_product"].apply(is_informative)).sum(),
            }), include_groups=False)
    except TypeError:
        pp = final_df.groupby("prophage").apply(
            lambda g: pd.Series({
                "total":         len(g),
                "hypo_in":       g["was_hypothetical"].sum(),
                "resolved":      (g["was_hypothetical"] & g["final_product"].apply(is_informative)).sum(),
                "still_hypo":    (g["was_hypothetical"] & ~g["final_product"].apply(is_informative)).sum(),
            }))
    log(pp.to_string())

    log("\nPartial structure matches (fs_qcov_frac < 0.5):")
    partial = final_df[final_df["fs_qcov_frac"].apply(
        lambda v: pd.notna(v) and float(v) < 0.5)]
    if len(partial):
        for _, r in partial.iterrows():
            log(f"  {r['locus_tag']}: {r['final_product']}  "
                f"(conf={r['fs_confidence']}, qcov={float(r['fs_qcov_frac']):.2f})")
    else:
        log("  none")

    # -- Update GenBank -------------------------------------------------------
    section("UPDATING GENBANK FILE")
    raw_gb = RAW_GB
    log(f"  Looking for source GBK: {raw_gb}")
    if not raw_gb.exists():
        # Cluster runs phold per-prophage; try to auto-concatenate from
        # PHAROKKA_OUT_DIR/<PROPHAGE>/<PROPHAGE>.gbk (one per prophage).
        try:
            from config import PHAROKKA_OUT_DIR
            import glob as _glob
            gbk_files = sorted(_glob.glob(str(PHAROKKA_OUT_DIR / "*" / "*.gbk")))
            log(f"  RAW_GB not found at {raw_gb}")
            log(f"  Scanning {PHAROKKA_OUT_DIR} for per-prophage GBKs: {len(gbk_files)} found")
            if gbk_files:
                raw_gb.parent.mkdir(parents=True, exist_ok=True)
                log(f"  Concatenating {len(gbk_files)} GBKs -> {raw_gb.name}")
                all_records = []
                for gf in gbk_files:
                    all_records.extend(SeqIO.parse(gf, "genbank"))
                with open(str(raw_gb), "w") as fh:
                    SeqIO.write(all_records, fh, "genbank")
                log(f"  Combined GBK written: {len(all_records)} records -> {raw_gb}")
            else:
                log(f"  No per-prophage GBKs found in {PHAROKKA_OUT_DIR}/*/*.gbk "
                    "(expected -- protein mode skips Pharokka)")
                log("  Trying pre-CDS-computed submode: build base GBK from "
                    "gene_metadata.csv + source genome + prophage window coords...")
                gbk_metadata = _enrich_metadata_for_gbk(metadata, RICH_GENE_METADATA_CSV)
                n_built = build_gbk_from_metadata(
                    gbk_metadata, SOURCE_GENOME_FASTA, PROPHAGE_WINDOWS_CSV,
                    PHOLD_AA_FASTA, raw_gb)
                if n_built == 0:
                    log("WARNING: Could not build a GBK (neither Pharokka GBKs nor "
                        "pre-CDS-computed metadata+genome were usable).")
                    log("  → Skipping GenBank update. CSV/XLSX outputs are complete.")
                    log("  → To enable for pre-CDS-computed datasets (e.g. citro), add:")
                    log(f"      1) source genome FASTA  -> {SOURCE_GENOME_FASTA}")
                    log(f"      2) window-coords CSV    -> {PROPHAGE_WINDOWS_CSV}  "
                        "(columns: prophage,genome_start,genome_stop)")
                    log("    ...or place a combined GBK at: " + str(RAW_GB))
                    log("  → Step 06 (phynteny) requires updated_prophages.gb — it will be "
                        "skipped until this is resolved.")
                    raw_gb = None
        except Exception as e:
            log(f"WARNING: Could not auto-build RAW_GB: {e}")
            log("  → Skipping GenBank update. CSV/XLSX outputs are complete.")
            raw_gb = None

    if raw_gb is not None and raw_gb.exists():
        log(f"  Source GB: {raw_gb}")
        n_updated = update_gb(raw_gb, annotation_map, UPDATED_GB)
        log(f"  Updated GB -> {UPDATED_GB}  ({n_updated} CDS updated)")
        updated_records = list(SeqIO.parse(str(UPDATED_GB), "genbank"))
        n_cds_gb = sum(1 for r in updated_records for f in r.features if f.type == "CDS")
        log(f"  CDS in updated GB: {n_cds_gb}  (expected: {total})")
        if n_cds_gb != total:
            log("  WARNING: CDS count mismatch -- check updated GB manually.")
        gb_status = f"{UPDATED_GB.name}  ({n_updated} CDS updated)"
    else:
        gb_status = "SKIPPED (no source GBK found — step 06 phynteny will not run)"

    section("PIPELINE COMPLETE")
    log(f"  {FINAL_ANNOTATIONS_TABLE.name}  ({total} genes, {len(final_df.columns)} cols)")
    log(f"  {FINAL_ANNOTATIONS_XLSX.name}   (Excel, colour-coded)")
    log(f"  {gb_status}")
    log(f"\n  Hypothetical resolution: {now_annotated}/{hypo_total} ({100*now_annotated//hypo_total}%)")
    log(f"  Location: {OUTPUT_DIR}/")


if __name__ == "__main__":
    main()
